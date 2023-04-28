from django.shortcuts import render
import openpyxl

# Create your views here.

#from django.shortcuts import render
from django.http import HttpResponse
from PyPDF2 import PdfReader
import re

def read_pdf(request):
    if request.method == 'POST':
        pdf_file = request.FILES['pdf_file']
        pdf_reader = PdfReader(pdf_file)
        contents = ''
        #pattern = r'CLIENT:\s*(.*?)\s*ORDERS:\s*(.*?)\s*WAVE:\s*(.*?)\s*Created:\s*(.*?)\s*'
        pattern2 =r'CLIENT: (.*?)\nORDERS: (.*?)\nWAVE: (.*?)\n.*?Created: (.*?) EDT'
        for page in pdf_reader.pages:
            contents += page.extract_text()
        matches = re.findall(pattern2, contents)
        output = ''
        for match in matches:
            output += f'Client: {match[0]}\nOrders: {match[1]}\nWave: {match[2]}\nCreated: {match[3]}\n\n<br>'
        
        #return render(request, 'result.html', {'matches': matches})
        #return HttpResponse(output)
        # create a new Excel workbook
        workbook = openpyxl.Workbook()        
        # select the active worksheet
        worksheet = workbook.active

        # write the headers to the worksheet
        worksheet.cell(row=1, column=1).value = "Client"
        worksheet.cell(row=1, column=2).value = "Orders"
        worksheet.cell(row=1, column=3).value = "Wave"
        worksheet.cell(row=1, column=4).value = "Created"
        worksheet.cell(row=1, column=5).value = "Assigned Employee"
        worksheet.cell(row=1, column=6).value = "Assigned time"
        worksheet.cell(row=1, column=7).value = "Pick Complete Time"

        # write the data to the worksheet
        for index, match in enumerate(matches, start=2):
            worksheet.cell(row=index, column=1).value = match[0]
            worksheet.cell(row=index, column=2).value = match[1]
            worksheet.cell(row=index, column=3).value = match[2]
            worksheet.cell(row=index, column=4).value = match[3]

        # save the workbook to a file
        workbook.save("output.xlsx")
        return render(request, 'result.html', {'matches': matches})
    else:
        return render(request, 'read_pdf.html')
    


